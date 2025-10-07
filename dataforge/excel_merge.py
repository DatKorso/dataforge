"""Excel file merging and filtering utilities.

This module provides clean, testable functions for merging multiple Excel files
with support for filtering by brand, article codes, and custom header configurations.
"""

from __future__ import annotations

import logging
import tempfile
from collections.abc import Callable
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


@dataclass
class SheetConfig:
    """Configuration for a single sheet in merge operation."""

    name: str
    include: bool = True
    header_rows: int = 0
    filter_by_brand: bool = False
    filter_by_articles: bool = False


@dataclass
class MergeConfig:
    """Complete configuration for Excel merge operation."""

    sheets: dict[str, SheetConfig]
    brand_filter: str | None = None
    brand_column_name: str = "Бренд в одежде и обуви"
    article_column_name: str = "Артикул"
    template_sheet_name: str = "Шаблон"
    video_sheet_names: list[str] | None = None
    # New options
    append_mode: bool = True  # If True, do not recreate sheets, only append rows after existing content
    filter_brand_at_end: bool = False  # If True, apply brand filter only after all appends

    def __post_init__(self):
        if self.video_sheet_names is None:
            self.video_sheet_names = ["Озон.Видео", "Озон.Видеообложка"]


class ExcelMergeError(Exception):
    """Base exception for Excel merge operations."""

    pass


class SheetNotFoundError(ExcelMergeError):
    """Raised when a required sheet is not found in workbook."""

    pass


class ColumnNotFoundError(ExcelMergeError):
    """Raised when a required column is not found in sheet."""

    pass


def find_column_index(
    df: pd.DataFrame, column_name: str, search_row: int = 1
) -> int | None:
    """Find column index by searching for column name in specified row.

    Args:
        df: DataFrame to search in
        column_name: Name or partial name to search for
        search_row: Row index to search in (0-based)

    Returns:
        Column index if found, None otherwise
    """
    if df.empty or search_row >= len(df):
        return None

    # Normalize search string for more robust matching
    normalized_search = column_name.lower().strip()

    for col_idx in range(df.shape[1]):
        try:
            cell_value = str(df.iat[search_row, col_idx]) if not pd.isna(df.iat[search_row, col_idx]) else ""
            # Normalize cell value: lowercase, strip whitespace, remove common artifacts
            normalized_cell = cell_value.lower().strip().replace("*", "").replace("\n", " ")
            
            # Try both substring match and exact match for flexibility
            if normalized_search in normalized_cell or normalized_cell in normalized_search:
                logger.debug(
                    f"Column match: '{column_name}' found in column {col_idx} "
                    f"(cell value: '{cell_value[:50]}...' at row {search_row})"
                )
                return col_idx
        except (IndexError, KeyError):
            continue

    # Log all header values if column not found for debugging
    logger.warning(
        f"Column '{column_name}' not found at row {search_row}. "
        f"Available values: {[str(df.iat[search_row, i])[:30] if not pd.isna(df.iat[search_row, i]) else 'NaN' for i in range(min(10, df.shape[1]))]}"
    )
    return None


def filter_by_brand(
    df: pd.DataFrame,
    brand_value: str,
    brand_column_index: int,
    header_rows: int = 4,
) -> pd.DataFrame:
    """Filter DataFrame rows by brand value.

    Args:
        df: DataFrame to filter
        brand_value: Brand value to filter by (case-insensitive substring match)
        brand_column_index: Index of brand column
        header_rows: Number of header rows to preserve

    Returns:
        Filtered DataFrame with headers and matching data rows
    """
    if df.empty or brand_column_index >= df.shape[1]:
        logger.warning(f"Cannot filter by brand: empty df or invalid column index {brand_column_index}")
        return df

    try:
        if len(df) <= header_rows:
            logger.debug(f"Brand filter: DataFrame has {len(df)} rows, not enough data beyond {header_rows} header rows")
            return df

        header = df.iloc[:header_rows].copy()
        data = df.iloc[header_rows:].copy()

        before_count = len(data)
        mask = (
            data.iloc[:, brand_column_index]
            .astype(str)
            .str.contains(brand_value, case=False, na=False)
        )
        filtered_data = data[mask]
        after_count = len(filtered_data)

        logger.info(
            f"Brand filter applied: column_idx={brand_column_index}, "
            f"brand='{brand_value}', rows {before_count}->{after_count}"
        )

        return pd.concat([header, filtered_data], ignore_index=True)
    except Exception as e:
        logger.error(f"Error filtering by brand: {e}", exc_info=True)
        return df


def extract_article_codes(
    file_bytes: bytes, sheet_name: str, article_column_name: str = "Артикул"
) -> set[str]:
    """Extract all article codes from a specific sheet.

    Args:
        file_bytes: Excel file as bytes
        sheet_name: Name of sheet to read
        article_column_name: Name/pattern of article column

    Returns:
        Set of article codes found
    """
    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(file_bytes)
            tmp.flush()
            tmp_path = Path(tmp.name)

        try:
            # Try reading with pandas first
            try:
                df = pd.read_excel(tmp_path, sheet_name=sheet_name, header=None)
            except Exception as parse_err:
                # Handle XML parsing errors
                if "not well-formed" in str(parse_err) or "ParseError" in str(type(parse_err).__name__):
                    logger.warning(
                        f"Sheet '{sheet_name}' has XML parsing error during article extraction: {parse_err}"
                    )
                    
                    df = None
                    # Try read_only mode
                    try:
                        logger.info(f"Attempting read_only recovery for article extraction from '{sheet_name}'...")
                        wb_repair = openpyxl.load_workbook(
                            tmp_path, read_only=True, data_only=True, keep_links=False
                        )
                        if sheet_name in wb_repair.sheetnames:
                            ws_repair = wb_repair[sheet_name]
                            data_rows = [list(row) for row in ws_repair.iter_rows(values_only=True)]
                            wb_repair.close()
                            if data_rows:
                                df = pd.DataFrame(data_rows)
                                logger.info(f"Successfully recovered sheet '{sheet_name}' for article extraction")
                    except Exception as repair_err:
                        logger.warning(f"Read-only recovery failed for '{sheet_name}': {repair_err}")
                    
                    if df is None:
                        logger.error(f"Failed to recover sheet '{sheet_name}' for article extraction, returning empty set")
                        return set()
                else:
                    raise

            if len(df) < 2:
                return set()

            # Find article column
            article_col_idx = find_column_index(df, article_column_name, search_row=1)
            if article_col_idx is None:
                logger.warning(f"Article column '{article_column_name}' not found in sheet '{sheet_name}'")
                return set()

            # Extract values starting from row 2 (after headers)
            article_codes = set()
            for row_idx in range(2, len(df)):
                if article_col_idx < df.shape[1]:
                    cell_value = df.iat[row_idx, article_col_idx]
                    if not pd.isna(cell_value):
                        code = str(cell_value).strip()
                        if code:
                            article_codes.add(code)

            logger.info(f"Extracted {len(article_codes)} article codes from {sheet_name}")
            return article_codes

        finally:
            tmp_path.unlink(missing_ok=True)

    except Exception as e:
        logger.error(f"Error extracting article codes from {sheet_name}: {e}", exc_info=True)
        return set()


def filter_by_articles(
    df: pd.DataFrame,
    article_codes: set[str],
    article_column_name: str = "Артикул",
    header_rows: int = 2,
) -> pd.DataFrame:
    """Filter DataFrame rows by article codes.

    Args:
        df: DataFrame to filter
        article_codes: Set of article codes to keep
        article_column_name: Name/pattern of article column
        header_rows: Number of header rows to preserve

    Returns:
        Filtered DataFrame with headers and matching data rows
    """
    if df.empty or not article_codes:
        return df

    try:
        if len(df) < header_rows:
            return df

        # Find article column
        article_col_idx = find_column_index(df, article_column_name, search_row=1)
        if article_col_idx is None:
            logger.warning(f"Article column '{article_column_name}' not found, returning headers only")
            return df.iloc[:header_rows] if len(df) >= header_rows else df

        header = df.iloc[:header_rows].copy()
        data = df.iloc[header_rows:].copy()

        # Filter by article codes
        filtered_rows = []
        for row_idx in range(len(data)):
            if article_col_idx < data.shape[1]:
                cell_value = data.iat[row_idx, article_col_idx]
                if not pd.isna(cell_value):
                    code = str(cell_value).strip()
                    if code in article_codes:
                        filtered_rows.append(data.iloc[row_idx : row_idx + 1])

        if filtered_rows:
            filtered_data = pd.concat(filtered_rows, ignore_index=False)
            return pd.concat([header, filtered_data], ignore_index=True)
        return header

    except Exception as e:
        logger.error(f"Error filtering by articles: {e}", exc_info=True)
        return df


def read_excel_sheet(
    file_bytes: bytes,
    sheet_name: str,
    config: MergeConfig,
    sheet_config: SheetConfig,
    template_articles: set[str] | None = None,
    apply_filters: bool = True,
) -> pd.DataFrame:
    """Read and filter a single Excel sheet according to configuration.

    Args:
        file_bytes: Excel file as bytes
        sheet_name: Name of sheet to read
        config: Global merge configuration
        sheet_config: Sheet-specific configuration
        template_articles: Optional article codes for filtering (for video sheets)

    Returns:
        Filtered DataFrame
    """
    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(file_bytes)
            tmp.flush()
            tmp_path = Path(tmp.name)

        try:
            # Try reading with pandas first
            try:
                df = pd.read_excel(tmp_path, sheet_name=sheet_name, header=None)
            except Exception as parse_err:
                # Handle XML parsing errors (corrupted Excel files)
                if "not well-formed" in str(parse_err) or "ParseError" in str(type(parse_err).__name__):
                    logger.warning(
                        f"Sheet '{sheet_name}' has XML parsing error: {parse_err}"
                    )
                    
                    # Try multiple recovery strategies
                    df = None
                    
                    # Strategy 1: Load with data_only (ignore formulas)
                    try:
                        logger.info(f"Attempting repair strategy 1 (data_only) for '{sheet_name}'...")
                        wb_repair = openpyxl.load_workbook(tmp_path, read_only=False, data_only=True)
                        if sheet_name in wb_repair.sheetnames:
                            ws_repair = wb_repair[sheet_name]
                            data_rows = []
                            for row in ws_repair.iter_rows(values_only=True):
                                data_rows.append(list(row))
                            wb_repair.close()
                            if data_rows:
                                df = pd.DataFrame(data_rows)
                                logger.info(f"Successfully recovered {len(df)} rows from '{sheet_name}' using strategy 1")
                        else:
                            logger.warning(f"Sheet '{sheet_name}' not found in repair workbook")
                    except Exception as repair1_err:
                        logger.warning(f"Strategy 1 failed for '{sheet_name}': {repair1_err}")
                    
                    # Strategy 2: Try read_only mode with keep_links=False
                    if df is None:
                        try:
                            logger.info(f"Attempting repair strategy 2 (read_only) for '{sheet_name}'...")
                            wb_repair = openpyxl.load_workbook(
                                tmp_path, read_only=True, data_only=True, keep_links=False
                            )
                            if sheet_name in wb_repair.sheetnames:
                                ws_repair = wb_repair[sheet_name]
                                data_rows = []
                                for row in ws_repair.iter_rows(values_only=True):
                                    data_rows.append(list(row))
                                wb_repair.close()
                                if data_rows:
                                    df = pd.DataFrame(data_rows)
                                    logger.info(f"Successfully recovered {len(df)} rows from '{sheet_name}' using strategy 2")
                        except Exception as repair2_err:
                            logger.warning(f"Strategy 2 failed for '{sheet_name}': {repair2_err}")
                    
                    # If all strategies failed, return empty DataFrame
                    if df is None:
                        logger.error(
                            f"All repair strategies failed for sheet '{sheet_name}'. "
                            f"This sheet will be skipped. Original error: {parse_err}"
                        )
                        return pd.DataFrame()
                else:
                    # Re-raise non-XML errors
                    raise

            if apply_filters:
                # Apply brand filter if configured
                if sheet_config.filter_by_brand and config.brand_filter:
                    brand_col_idx = find_column_index(df, config.brand_column_name, search_row=1)
                    if brand_col_idx is not None:
                        df = filter_by_brand(
                            df, config.brand_filter, brand_col_idx, header_rows=sheet_config.header_rows
                        )
                    else:
                        logger.warning(f"Brand column not found in sheet '{sheet_name}', skipping brand filter")

                # Apply article filter if configured (except defer for video sheets to final stage)
                if (
                    sheet_config.filter_by_articles
                    and template_articles
                    and sheet_name not in (config.video_sheet_names or [])
                ):
                    df = filter_by_articles(
                        df,
                        template_articles,
                        article_column_name=config.article_column_name,
                        header_rows=sheet_config.header_rows,
                    )

            return df

        finally:
            tmp_path.unlink(missing_ok=True)

    except Exception as e:
        logger.error(f"Error reading sheet '{sheet_name}': {e}", exc_info=True)
        return pd.DataFrame()


def merge_excel_files(
    template_bytes: bytes,
    additional_files_bytes: list[bytes],
    config: MergeConfig,
    progress_callback: Callable[[float, str], None] | None = None,
) -> bytes:
    """Merge multiple Excel files according to configuration.

    Args:
        template_bytes: Template Excel file as bytes
        additional_files_bytes: List of additional files to merge
        config: Merge configuration
        progress_callback: Optional callback for progress updates (progress: float, message: str)

    Returns:
        Merged Excel file as bytes

    Raises:
        ExcelMergeError: If merge operation fails
    """
    tmp_template_path: Path | None = None

    def report_progress(pct: float, msg: str):
        if progress_callback:
            try:
                progress_callback(pct, msg)
            except Exception as e:
                logger.warning(f"Progress callback error: {e}")

    try:
        report_progress(0.05, "🔍 Инициализация слияния...")

        # Extract article codes from template if needed for video sheets
        template_articles: set[str] | None = None
        video_sheets_to_process = [
            name
            for name, sheet_cfg in config.sheets.items()
            if sheet_cfg.include and name in (config.video_sheet_names or [])
        ]

        if video_sheets_to_process:
            report_progress(0.10, "📋 Извлечение артикулов из шаблона...")
            template_articles = extract_article_codes(
                template_bytes,
                config.template_sheet_name,
                article_column_name=config.article_column_name,
            )

        # Create temp file for template
        report_progress(0.15, "📂 Загрузка шаблона...")
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False, mode="wb") as tmp:
            tmp.write(template_bytes)
            tmp.flush()
            tmp_template_path = Path(tmp.name)

        wb = openpyxl.load_workbook(tmp_template_path)
        all_sheets = wb.sheetnames

        total_sheets = len([s for s in config.sheets.values() if s.include])
        processed = 0

        # Process each sheet
        for sheet_name in all_sheets:
            sheet_cfg = config.sheets.get(sheet_name)
            
            # Skip sheets that are not configured or not included for merging
            # (they will remain in the workbook unchanged)
            if not sheet_cfg or not sheet_cfg.include:
                logger.info(f"Skipping sheet '{sheet_name}' (not included in merge)")
                continue

            processed += 1
            progress = 0.15 + (0.7 * processed / total_sheets)
            report_progress(progress, f"📋 Объединение листа '{sheet_name}'...")

            # Read template sheet with filters
            # In append_mode we do not want to filter template sheet early if brand filter postponed
            template_df = read_excel_sheet(
                template_bytes,
                sheet_name,
                config,
                sheet_cfg,
                template_articles=template_articles,
                apply_filters=not config.filter_brand_at_end,
            )
            
            # If template sheet is corrupted and returned empty, skip merging for this sheet
            if template_df.empty:
                logger.warning(
                    f"Template sheet '{sheet_name}' is empty or corrupted, skipping merge for this sheet. "
                    f"Original content will be preserved."
                )
                continue

            # Read and merge additional files
            additional_dfs = []
            for file_bytes in additional_files_bytes:
                df = read_excel_sheet(
                    file_bytes,
                    sheet_name,
                    config,
                    sheet_cfg,
                    template_articles=template_articles,
                    apply_filters=not config.filter_brand_at_end,
                )
                if not df.empty and len(df) > sheet_cfg.header_rows:
                    df = df.iloc[sheet_cfg.header_rows :]
                    additional_dfs.append(df)

            ws = wb[sheet_name]

            if config.append_mode:
                # Keep existing data; only append additional rows (already without headers)
                # Important: if we're filtering early (not deferred), we need to replace
                # the existing template data with filtered version first
                if not config.filter_brand_at_end and (sheet_cfg.filter_by_brand and config.brand_filter):
                    # Replace existing template data with filtered version
                    logger.info(f"Applying early brand filter to template in append mode for '{sheet_name}'")
                    ws.delete_rows(1, ws.max_row)
                    for r_idx, row in enumerate(
                        dataframe_to_rows(template_df, index=False, header=False)
                    ):
                        for c_idx, value in enumerate(row):
                            ws.cell(row=r_idx + 1, column=c_idx + 1, value=value)
                elif ws.max_row == 0 or ws.max_row == 1 and all([cell.value is None for cell in ws[1]]):
                    # Write whole template_df as baseline if sheet is empty
                    for r_idx, row in enumerate(
                        dataframe_to_rows(template_df, index=False, header=False)
                    ):
                        for c_idx, value in enumerate(row):
                            ws.cell(row=r_idx + 1, column=c_idx + 1, value=value)
                
                # Append additional rows at end
                start_row = ws.max_row + 1
                for add_df in additional_dfs:
                    for row in dataframe_to_rows(add_df, index=False, header=False):
                        if all(v is None for v in row):
                            continue
                        for c_idx, value in enumerate(row):
                            ws.cell(row=start_row, column=c_idx + 1, value=value)
                        start_row += 1
            else:
                # Rebuild sheet from scratch
                combined_df = (
                    pd.concat([template_df] + additional_dfs, ignore_index=True)
                    if additional_dfs
                    else template_df
                )
                ws.delete_rows(1, ws.max_row)
                for r_idx, row in enumerate(
                    dataframe_to_rows(combined_df, index=False, header=False)
                ):
                    for c_idx, value in enumerate(row):
                        ws.cell(row=r_idx + 1, column=c_idx + 1, value=value)

        report_progress(0.90, "💾 Сохранение результата...")

        # Final deferred filtering phase (brand and video sheet article filters)
        if config.filter_brand_at_end or any(
            cfg.filter_by_articles and name in (config.video_sheet_names or []) for name, cfg in config.sheets.items()
        ):
            report_progress(0.92, "🧹 Финальные фильтры...")

            # Step 1: First pass - apply brand filter to template and other sheets
            # This ensures we get the correct article list from filtered template
            if config.filter_brand_at_end and config.brand_filter:
                for sheet_name, sheet_cfg in config.sheets.items():
                    if not sheet_cfg.include or not sheet_cfg.filter_by_brand:
                        continue
                    
                    # Skip video sheets in this pass - they need article filter after brand
                    is_video_sheet = sheet_name in (config.video_sheet_names or [])
                    if is_video_sheet and sheet_cfg.filter_by_articles:
                        continue
                    
                    try:
                        ws = wb[sheet_name]
                    except KeyError:
                        logger.warning(f"Sheet '{sheet_name}' not found during brand filtering")
                        continue

                    try:
                        rows_data = [list(r) for r in ws.iter_rows(values_only=True)]
                    except Exception as iter_err:
                        logger.error(f"Failed to read sheet '{sheet_name}' for brand filtering: {iter_err}")
                        continue
                    
                    if not rows_data:
                        continue
                    
                    df_sheet = pd.DataFrame(rows_data)
                    brand_col_idx = find_column_index(df_sheet, config.brand_column_name, search_row=1)
                    
                    if brand_col_idx is not None:
                        before_count = len(df_sheet)
                        df_sheet = filter_by_brand(
                            df_sheet,
                            config.brand_filter,
                            brand_col_idx,
                            header_rows=sheet_cfg.header_rows,
                        )
                        after_count = len(df_sheet)
                        logger.info(f"Brand filter for '{sheet_name}': {before_count}->{after_count} rows")
                        
                        # Rewrite sheet
                        ws.delete_rows(1, ws.max_row)
                        for r_idx, row in enumerate(dataframe_to_rows(df_sheet, index=False, header=False)):
                            for c_idx, value in enumerate(row):
                                ws.cell(row=r_idx + 1, column=c_idx + 1, value=value)
                    else:
                        logger.warning(f"Brand column not found in '{sheet_name}'")

            # Step 2: Extract article codes from NOW FILTERED template sheet
            final_template_articles: set[str] | None = None
            need_video_articles = any(
                cfg.filter_by_articles and name in (config.video_sheet_names or [])
                for name, cfg in config.sheets.items()
                if cfg.include
            )
            
            if need_video_articles:
                try:
                    tmpl_ws = wb[config.template_sheet_name]
                    try:
                        data = [list(r) for r in tmpl_ws.iter_rows(values_only=True)]
                    except Exception as iter_err:
                        logger.error(
                            f"Failed to iterate rows in template sheet '{config.template_sheet_name}': {iter_err}. "
                            f"Skipping final article extraction.",
                            exc_info=True
                        )
                        data = []
                    
                    if data:
                        df_template = pd.DataFrame(data)
                        art_idx = find_column_index(df_template, config.article_column_name, search_row=1)
                        if art_idx is not None and len(df_template) > 2:
                            codes = set()
                            for row_idx in range(2, len(df_template)):
                                val = df_template.iat[row_idx, art_idx]
                                if not pd.isna(val):
                                    s = str(val).strip()
                                    if s:
                                        codes.add(s)
                            final_template_articles = codes
                            logger.info(
                                f"Article codes extracted from FILTERED template: {len(final_template_articles)} items"
                            )
                        else:
                            logger.warning("Could not find article column in filtered template")
                except Exception as e:
                    logger.error(f"Failed to extract articles from filtered template: {e}", exc_info=True)

            # Step 3: Process video sheets with brand + article filters
            for sheet_name, sheet_cfg in config.sheets.items():
                if not sheet_cfg.include:
                    continue
                
                is_video_sheet = sheet_name in (config.video_sheet_names or [])
                
                # Only process video sheets with article filter in this pass
                if not (is_video_sheet and sheet_cfg.filter_by_articles):
                    continue
                
                try:
                    ws = wb[sheet_name]
                except KeyError:
                    logger.warning(f"Video sheet '{sheet_name}' not found")
                    continue

                try:
                    rows_data = [list(r) for r in ws.iter_rows(values_only=True)]
                except Exception as iter_err:
                    logger.error(f"Failed to read video sheet '{sheet_name}': {iter_err}")
                    continue
                
                if not rows_data:
                    continue
                
                df_sheet = pd.DataFrame(rows_data)
                original_count = len(df_sheet)

                # Apply brand filter first
                if config.brand_filter and sheet_cfg.filter_by_brand:
                    brand_col_idx = find_column_index(df_sheet, config.brand_column_name, search_row=1)
                    if brand_col_idx is not None:
                        df_sheet = filter_by_brand(
                            df_sheet,
                            config.brand_filter,
                            brand_col_idx,
                            header_rows=sheet_cfg.header_rows,
                        )
                        after_brand = len(df_sheet)
                        logger.info(f"Video sheet '{sheet_name}' brand filter: {original_count}->{after_brand} rows")
                    else:
                        logger.warning(f"Brand column not found in video sheet '{sheet_name}'")

                # Apply article filter second (using filtered template articles)
                if final_template_articles:
                    before_article = len(df_sheet)
                    df_sheet = filter_by_articles(
                        df_sheet,
                        final_template_articles,
                        article_column_name=config.article_column_name,
                        header_rows=sheet_cfg.header_rows,
                    )
                    after_article = len(df_sheet)
                    logger.info(f"Video sheet '{sheet_name}' article filter: {before_article}->{after_article} rows")

                # Rewrite sheet
                ws.delete_rows(1, ws.max_row)
                for r_idx, row in enumerate(dataframe_to_rows(df_sheet, index=False, header=False)):
                    for c_idx, value in enumerate(row):
                        ws.cell(row=r_idx + 1, column=c_idx + 1, value=value)

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        wb.close()

        report_progress(1.0, "✅ Готово!")

        return output.getvalue()

    except Exception as e:
        logger.error(f"Error merging Excel files: {e}", exc_info=True)
        raise ExcelMergeError(f"Не удалось объединить файлы: {e}") from e

    finally:
        if tmp_template_path and tmp_template_path.exists():
            try:
                tmp_template_path.unlink()
            except Exception as e:
                logger.warning(f"Failed to cleanup temp file: {e}")


def get_sheet_info(file_bytes: bytes) -> list[tuple[str, int]]:
    """Get list of sheets with row counts from Excel file.

    Args:
        file_bytes: Excel file as bytes

    Returns:
        List of (sheet_name, row_count) tuples
    """
    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(file_bytes)
            tmp.flush()
            tmp_path = Path(tmp.name)

        try:
            wb = openpyxl.load_workbook(tmp_path, read_only=True)
            result = []
            for name in wb.sheetnames:
                ws = wb[name]
                max_row = ws.max_row or 0
                result.append((name, max_row))
            wb.close()
            return result

        finally:
            tmp_path.unlink(missing_ok=True)

    except Exception as e:
        logger.error(f"Error getting sheet info: {e}", exc_info=True)
        return []
