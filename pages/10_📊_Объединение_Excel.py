from __future__ import annotations

import io
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl
import pandas as pd
import os
import tempfile
import traceback
from io import BytesIO
from openpyxl.utils.dataframe import dataframe_to_rows
import streamlit as st

from dataforge.ui import guard_page, setup_page


setup_page(title="Объединение эксель", icon="📎")
guard_page("enable_excel_merge", default=True, message="Страница объединения Excel отключена.")

st.title("📎 Объединение эксель")
st.caption(
    "Выберите исходный файл (.xlsx), укажите число строк заголовка для каждого листа, затем добавьте другие файлы для объединения."
)


def _list_sheets_and_counts(wb: openpyxl.Workbook) -> List[Tuple[str, int]]:
    out: List[Tuple[str, int]] = []
    for name in wb.sheetnames:
        ws = wb[name]
        # approximate last used row
        max_row = ws.max_row or 0
        out.append((name, max_row))
    return out


def _copy_sheet_to_target(
    src_ws,
    tgt_ws,
    header_rows: int,
    is_initial: bool,
) -> int:
    """Copy rows from src_ws into tgt_ws. If is_initial True, copy all rows; otherwise skip header_rows and append remaining.

    Returns number of rows appended (excluding headers for non-initial files).
    """
    rows_copied = 0
    start_row = 1 if is_initial else (header_rows + 1)

    # detect write-only target (streaming) — WriteOnlyWorksheet supports append() but not cell()/max_row
    is_streaming_target = hasattr(tgt_ws, "append") and not hasattr(tgt_ws, "cell")

    if is_streaming_target:
        # faster, memory-efficient path: append row values only (styles lost)
        for values in src_ws.iter_rows(min_row=start_row, max_row=src_ws.max_row, values_only=True):
            # values is a tuple; convert to list to ensure append accepts it
            try:
                tgt_ws.append(list(values))
            except Exception:
                # fallback: try appending tuple
                try:
                    tgt_ws.append(values)
                except Exception:
                    pass
            rows_copied += 1
        return rows_copied

    # normal (in-memory) path: copy cell-by-cell and attempt to preserve styles
    for r_idx, row in enumerate(src_ws.iter_rows(min_row=start_row, max_row=src_ws.max_row), start=1):
        tgt_row_idx = tgt_ws.max_row + 1
        for c_idx, cell in enumerate(row, start=1):
            tgt_cell = tgt_ws.cell(row=tgt_row_idx, column=c_idx, value=cell.value)
            # copy simple style attributes if present
            if getattr(cell, "has_style", False):
                try:
                    tgt_cell.font = cell.font
                    tgt_cell.fill = cell.fill
                    tgt_cell.border = cell.border
                    tgt_cell.alignment = cell.alignment
                    tgt_cell.number_format = cell.number_format
                except Exception:
                    # be forgiving if styles can't be copied
                    pass
        rows_copied += 1

    # copy merged cell contents as plain cells: find merged ranges in src and set value in corresponding target cell(s)
    try:
        for merged in list(getattr(src_ws, "merged_cells").ranges):
            min_col, min_row, max_col, max_row = merged.bounds
            src_val = src_ws.cell(row=min_row, column=min_col).value
            # compute approx target row: append mode so target top-left will be shifted by difference in max_row
            # This is an approximation — exact mapping is complex when sheets have different sizes
            try:
                tgt_ws.cell(row=min_row + (tgt_ws.max_row - src_ws.max_row), column=min_col, value=src_val)
            except Exception:
                pass
    except Exception:
        # Non-critical
        pass

    return rows_copied


# --- Helpers borrowed from ported implementation ---
def get_excel_sheets(file_bytes: bytes) -> List[str]:
    try:
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            tmp_file.write(file_bytes)
            tmp_file.flush()

            wb = openpyxl.load_workbook(tmp_file.name, read_only=True)
            sheets = wb.sheetnames
            wb.close()
            os.unlink(tmp_file.name)
            return sheets
    except Exception:
        return []


def read_excel_sheet(file_bytes: bytes, sheet_name: str, start_row: int = 0) -> pd.DataFrame:
    try:
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            tmp_file.write(file_bytes)
            tmp_file.flush()

            df = pd.read_excel(tmp_file.name, sheet_name=sheet_name, header=None)
            os.unlink(tmp_file.name)

            if start_row > 0 and len(df) > start_row:
                df = df.iloc[start_row:]

            return df
    except Exception:
        return pd.DataFrame()


def filter_dataframe_by_brand(df: pd.DataFrame, brand_filter: str, brand_column_index: int | None, keep_headers: bool = True) -> pd.DataFrame:
    try:
        if df.empty or not brand_filter or brand_column_index is None:
            return df

        if keep_headers and len(df) > 4:
            header_rows = df.iloc[:4].copy()
            data_rows = df.iloc[4:].copy()
            if brand_column_index < len(data_rows.columns):
                brand_mask = data_rows.iloc[:, brand_column_index].astype(str).str.contains(brand_filter, case=False, na=False)
                filtered_data = data_rows[brand_mask]
                return pd.concat([header_rows, filtered_data], ignore_index=True)
            else:
                return header_rows
        else:
            if brand_column_index < len(df.columns):
                brand_mask = df.iloc[:, brand_column_index].astype(str).str.contains(brand_filter, case=False, na=False)
                return df[brand_mask]
            else:
                return pd.DataFrame()
    except Exception:
        return df


def get_template_article_values(template_bytes: bytes) -> set:
    try:
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            tmp_file.write(template_bytes)
            tmp_file.flush()

            df = pd.read_excel(tmp_file.name, sheet_name="Шаблон", header=None)
            os.unlink(tmp_file.name)

            if len(df) < 2:
                return set()

            article_column_index = None
            # iterate by integer column positions to avoid label/index ambiguity
            for col_idx in range(df.shape[1]):
                cell_value = str(df.iat[1, col_idx]) if not pd.isna(df.iat[1, col_idx]) else ""
                if "Артикул" in cell_value and "*" in cell_value:
                    article_column_index = col_idx
                    break

            if article_column_index is None:
                return set()

            article_values = set()
            for row_idx in range(2, len(df)):
                if article_column_index is not None and article_column_index < df.shape[1]:
                    cell_value = df.iat[row_idx, article_column_index]
                    if not pd.isna(cell_value):
                        article_value = str(cell_value).strip()
                        if article_value:
                            article_values.add(article_value)

            return article_values
    except Exception:
        return set()


def filter_dataframe_by_template_articles(df: pd.DataFrame, template_articles: set, sheet_name: str) -> pd.DataFrame:
    try:
        if df.empty or not template_articles:
            return df

        article_column_index = None
        if len(df) >= 2:
            for col_idx in range(df.shape[1]):
                cell_value = str(df.iat[1, col_idx]) if not pd.isna(df.iat[1, col_idx]) else ""
                if "Артикул" in cell_value and "*" in cell_value:
                    article_column_index = col_idx
                    break

        if article_column_index is None:
            return df.iloc[:2] if len(df) >= 2 else df

        if len(df) <= 2:
            return df

        header_rows = df.iloc[:2].copy()
        data_rows = df.iloc[2:].copy()

        filtered_data_list = []
        for row_idx in range(len(data_rows)):
            if article_column_index is not None and article_column_index < data_rows.shape[1]:
                cell_value = data_rows.iat[row_idx, article_column_index]
                if not pd.isna(cell_value):
                    article_value = str(cell_value).strip()
                    if article_value in template_articles:
                        filtered_data_list.append(data_rows.iloc[row_idx:row_idx+1])

        if filtered_data_list:
            filtered_data = pd.concat(filtered_data_list, ignore_index=False)
            return pd.concat([header_rows, filtered_data], ignore_index=True)
        else:
            return header_rows
    except Exception:
        return df


def read_and_filter_excel_sheet(file_bytes: bytes, sheet_name: str, start_row: int = 0, brand_filter: str | None = None, brand_column_index: int | None = None) -> pd.DataFrame:
    try:
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            tmp_file.write(file_bytes)
            tmp_file.flush()

            df = pd.read_excel(tmp_file.name, sheet_name=sheet_name, header=None)
            os.unlink(tmp_file.name)

            if brand_filter and brand_column_index is not None:
                df = filter_dataframe_by_brand(df, brand_filter, brand_column_index, keep_headers=False)

            if start_row > 0 and len(df) > start_row:
                df = df.iloc[start_row:]

            return df
    except Exception:
        return pd.DataFrame()


def check_brand_column_exists(file_bytes: bytes, sheet_name: str = "Шаблон") -> Tuple[bool, int | None]:
    try:
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            tmp_file.write(file_bytes)
            tmp_file.flush()

            df = pd.read_excel(tmp_file.name, sheet_name=sheet_name, header=None, skiprows=1, nrows=1)
            os.unlink(tmp_file.name)

            for col_idx in range(df.shape[1]):
                cell_value = str(df.iat[0, col_idx]) if not pd.isna(df.iat[0, col_idx]) else ""
                if "Бренд в одежде и обуви" in cell_value:
                    return True, col_idx
            return False, None
    except Exception:
        return False, None


def merge_excel_files(template_bytes: bytes, additional_files_bytes: List[bytes], sheet_config: Dict[str, Dict], brand_filter: str | None = None, progress_callback=None) -> bytes:
    """Optimized pandas-based merge. Returns bytes of resulting xlsx."""
    tmp_template = None
    try:
        if progress_callback:
            progress_callback(0.05, "🔍 Анализ конфигурации и проверка фильтрации...")

        brand_column_index = None
        has_brand_column = False

        if brand_filter and sheet_config.get("Шаблон", {}).get('merge'):
            if progress_callback:
                progress_callback(0.1, "🔍 Проверка наличия колонки бренда...")
            has_brand_column, brand_column_index = check_brand_column_exists(template_bytes, "Шаблон")
            if not has_brand_column:
                brand_filter = None

        template_articles = set()
        video_sheets = ["Озон.Видео", "Озон.Видеообложка"]
        video_sheets_to_process = [name for name, config in sheet_config.items() if config.get('merge') and name in video_sheets]
        if video_sheets_to_process:
            if progress_callback:
                progress_callback(0.13, "🔍 Извлечение артикулов из шаблона для фильтрации видео-листов...")
            template_articles = get_template_article_values(template_bytes)

        merge_sheets = [name for name, config in sheet_config.items() if config.get('merge')]

        # create temp file for template
        tmp_template = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False, mode='wb')
        tmp_template.write(template_bytes)
        tmp_template.close()
        template_wb = openpyxl.load_workbook(tmp_template.name)

        all_template_sheets = template_wb.sheetnames

        # For each sheet
        for sheet_name in all_template_sheets:
            if sheet_name in sheet_config and sheet_config[sheet_name].get('merge'):
                config = sheet_config[sheet_name]
                start_row = config.get('start_row', 0)
                if progress_callback:
                    progress_callback(0.15, f"📋 Объединение листа '{sheet_name}'...")

                # read and filter template sheet accordingly
                if sheet_name == "Шаблон" and brand_filter and has_brand_column:
                    template_df = pd.read_excel(tmp_template.name, sheet_name=sheet_name, header=None)
                    template_df = filter_dataframe_by_brand(template_df, brand_filter, brand_column_index, keep_headers=True)
                elif sheet_name in video_sheets and template_articles:
                    template_df = pd.read_excel(tmp_template.name, sheet_name=sheet_name, header=None)
                    template_df = filter_dataframe_by_template_articles(template_df, template_articles, sheet_name)
                else:
                    template_df = pd.read_excel(tmp_template.name, sheet_name=sheet_name, header=None)

                additional_data = []
                for add_file_bytes in additional_files_bytes:
                    add_df = read_and_filter_excel_sheet(add_file_bytes, sheet_name, start_row, brand_filter, brand_column_index)
                    if sheet_name in video_sheets and template_articles and not add_df.empty:
                        add_df = filter_dataframe_by_template_articles(add_df, template_articles, sheet_name)
                    if not add_df.empty:
                        additional_data.append(add_df)

                if additional_data:
                    combined_df = pd.concat([template_df] + additional_data, ignore_index=True)
                    ws = template_wb[sheet_name]
                    ws.delete_rows(1, ws.max_row)
                    for r_idx, row in enumerate(dataframe_to_rows(combined_df, index=False, header=False)):
                        for c_idx, value in enumerate(row):
                            ws.cell(row=r_idx + 1, column=c_idx + 1, value=value)
                else:
                    ws = template_wb[sheet_name]
                    ws.delete_rows(1, ws.max_row)
                    for r_idx, row in enumerate(dataframe_to_rows(template_df, index=False, header=False)):
                        for c_idx, value in enumerate(row):
                            ws.cell(row=r_idx + 1, column=c_idx + 1, value=value)

        output = BytesIO()
        template_wb.save(output)
        output.seek(0)
        template_wb.close()
        return output.getvalue()
    except Exception as e:
        raise
    finally:
        if tmp_template and hasattr(tmp_template, 'name') and os.path.exists(tmp_template.name):
            try:
                os.unlink(tmp_template.name)
            except Exception:
                pass

# --- End helpers ---


st.sidebar.header("Шаги")
st.sidebar.markdown("1) Выберите начальный .xlsx\n2) Укажите число строк заголовка для листа(ов)\n3) Добавьте другие .xlsx файлы и нажмите 'Объединить'.")

uploaded_initial = st.file_uploader("Выберите начальный .xlsx", type=["xlsx"], key="merge_initial")

if uploaded_initial is not None:
    try:
        in_mem = io.BytesIO(uploaded_initial.read())
        wb_initial = openpyxl.load_workbook(in_mem, data_only=False)
    except Exception as exc:
        st.error(f"Не удалось прочитать начальный файл: {exc}")
        st.stop()

    sheets = _list_sheets_and_counts(wb_initial)
    st.subheader("Листы начального файла")
    header_rows_map: Dict[str, int] = {}
    include_map: Dict[str, bool] = {}

    # Preset defaults
    PRESETS = {
        "Шаблон": 4,
        "Озон.Видео": 3,
        "Озон.Видеообложка": 3,
    }

    for name, max_row in sheets:
        # default include only for preset names
        default_include = name in PRESETS
        cols = st.columns([1, 1, 1])
        with cols[0]:
            include = st.checkbox(f"Включить '{name}'", value=default_include, key=f"inc_{name}")
            include_map[name] = include
        with cols[1]:
            cols[1].write(f"{name} (строк: {max_row})")
        with cols[2]:
            default_hdr = PRESETS.get(name, 1)
            # number_input disabled when sheet is not included
            header_rows_map[name] = st.number_input(
                f"Заголовков в '{name}'", min_value=0, value=int(default_hdr), step=1, key=f"hdr_{name}",
                disabled=not include_map[name],
            )

    st.markdown("---")

    st.markdown("---")
    uploaded_others = st.file_uploader(
        "Добавьте файлы для объединения (можно несколько)", type=["xlsx"], accept_multiple_files=True, key="merge_others"
    )

    st.sidebar.markdown("---")
    fast_mode = st.sidebar.checkbox(
        "Fast mode (streaming, меньше памяти, быстрее)", value=True,
        help="При включении используется write_only/read_only режим openpyxl: быстрее и экономит память, но некоторые стили/формулы могут быть потеряны"
    )

    progress_bar = st.progress(0)
    status = st.empty()

    if st.button("Объединить"):
        files_to_merge: List[Tuple[str, io.BytesIO]] = []
        # include initial as first
        in_mem.seek(0)
        files_to_merge.append((uploaded_initial.name, io.BytesIO(in_mem.read())))

        if uploaded_others:
            for f in uploaded_others:
                files_to_merge.append((f.name, io.BytesIO(f.read())))

        total_files = len(files_to_merge)
        if total_files == 0:
            st.warning("Нет файлов для объединения")
            st.stop()

        # If not fast_mode, use pandas-based optimized merge from helpers (supports filtering)
        if not fast_mode:
            try:
                template_bytes = files_to_merge[0][1].getvalue()
                additional_bytes = [b.getvalue() for (_, b) in files_to_merge[1:]]

                # build sheet_config from UI maps
                sheet_config_map: Dict[str, Dict] = {}
                for name, _ in sheets:
                    sheet_config_map[name] = {
                        'merge': bool(include_map.get(name, False)),
                        'start_row': int(header_rows_map.get(name, 0) or 0),
                    }

                # progress callback
                def cb(value, message=None):
                    try:
                        if message:
                            status.text(message)
                        progress_bar.progress(min(1.0, float(value)))
                    except Exception:
                        pass

                # apply default brand filter for Шаблон
                result_bytes = merge_excel_files(template_bytes, additional_bytes, sheet_config_map, brand_filter="Shuzzi", progress_callback=cb)
                if result_bytes:
                    st.success("Объединение завершено")
                    st.download_button("Скачать объединённый .xlsx", data=result_bytes, file_name="merged.xlsx")
                else:
                    st.error("Ошибка при объединении (пустой результат)")
            except Exception as exc:
                st.error(f"Ошибка при объединении: {exc}")
                with st.expander("Подробности ошибки"):
                    st.code(traceback.format_exc())

        else:
            # create target workbook in streaming mode
            tgt_wb = openpyxl.Workbook(write_only=True)

            files_processed = 0
            for idx, (fname, bio) in enumerate(files_to_merge):
                bio.seek(0)
                try:
                    src_wb = openpyxl.load_workbook(bio, data_only=False)
                except Exception as exc:
                    st.error(f"Не удалось открыть {fname}: {exc}")
                    continue

                for sheet_name in src_wb.sheetnames:
                    include = include_map.get(sheet_name, False)
                    if not include:
                        continue

                    src_ws = src_wb[sheet_name]
                    existing = [s for s in tgt_wb.worksheets if s.title == sheet_name]
                    if existing:
                        tgt_ws = existing[0]
                    else:
                        tgt_ws = tgt_wb.create_sheet(title=sheet_name)

                    is_initial = idx == 0
                    hdr = header_rows_map.get(sheet_name, 0)
                    status.text(f"Обработка {fname} — лист {sheet_name}")
                    _copy_sheet_to_target(src_ws, tgt_ws, int(hdr), is_initial)

                files_processed += 1
                progress = int(files_processed / total_files * 100)
                progress_bar.progress(progress)

            status.text("Готово — формирование файла для скачивания")
            out_stream = io.BytesIO()
            tgt_wb.save(out_stream)
            out_stream.seek(0)
            # Post-process: filter 'Шаблон' to keep only brand 'Shuzzi'
            def _filter_shablon_brand_bytes(in_bytes: bytes, header_rows: int = 4, brand_value: str = "Shuzzi") -> bytes:
                tmpf = None
                try:
                    tmpf = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False, mode='wb')
                    tmpf.write(in_bytes)
                    tmpf.close()

                    try:
                        df = pd.read_excel(tmpf.name, sheet_name="Шаблон", header=None)
                    except Exception:
                        return in_bytes

                    # find brand column in second row (index 1)
                    brand_col = None
                    if df.shape[0] >= 2:
                        for col_idx in range(df.shape[1]):
                            cell_value = str(df.iat[1, col_idx]) if not pd.isna(df.iat[1, col_idx]) else ""
                            if "Бренд в одежде и обуви" in cell_value:
                                brand_col = col_idx
                                break

                    if brand_col is None:
                        return in_bytes

                    header = df.iloc[:header_rows].copy() if df.shape[0] > header_rows else df.iloc[:header_rows].copy()
                    data = df.iloc[header_rows:].copy() if df.shape[0] > header_rows else pd.DataFrame()

                    if not data.empty and brand_col < data.shape[1]:
                        mask = data.iloc[:, brand_col].astype(str).str.contains(brand_value, case=False, na=False)
                        filtered = data[mask]
                    else:
                        filtered = pd.DataFrame()

                    combined = pd.concat([header, filtered], ignore_index=True)

                    # write back into workbook
                    wb = openpyxl.load_workbook(tmpf.name)
                    if "Шаблон" not in wb.sheetnames:
                        return in_bytes
                    ws = wb["Шаблон"]
                    ws.delete_rows(1, ws.max_row)
                    for r_idx, row in enumerate(dataframe_to_rows(combined, index=False, header=False)):
                        for c_idx, value in enumerate(row):
                            ws.cell(row=r_idx + 1, column=c_idx + 1, value=value)

                    outb = BytesIO()
                    wb.save(outb)
                    outb.seek(0)
                    wb.close()
                    return outb.getvalue()
                finally:
                    if tmpf and hasattr(tmpf, 'name') and os.path.exists(tmpf.name):
                        try:
                            os.unlink(tmpf.name)
                        except Exception:
                            pass

            final_bytes = _filter_shablon_brand_bytes(out_stream.getvalue(), header_rows=header_rows_map.get("Шаблон", 4), brand_value="Shuzzi")

            st.success("Объединение завершено")
            st.download_button("Скачать объединённый .xlsx", data=final_bytes, file_name="merged.xlsx")