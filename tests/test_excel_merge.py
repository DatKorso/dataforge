"""Tests for excel_merge module."""

from __future__ import annotations

import io

import openpyxl
import pandas as pd
import pytest
from dataforge.excel_merge import (
    ExcelMergeError,
    MergeConfig,
    SheetConfig,
    extract_article_codes,
    filter_by_articles,
    filter_by_brand,
    find_column_index,
    get_sheet_info,
    merge_excel_files,
    read_excel_sheet,
)


@pytest.fixture
def sample_dataframe():
    """Create a sample DataFrame with headers and data."""
    data = {
        0: ["Header1", "Артикул *", "ART001", "ART002", "ART003"],
        1: ["Header2", "Название", "Товар 1", "Товар 2", "Товар 3"],
        2: ["Header3", "Бренд в одежде и обуви", "Shuzzi", "OtherBrand", "Shuzzi"],
        3: ["Header4", "Цена", "100", "200", "300"],
    }
    return pd.DataFrame(data)


@pytest.fixture
def sample_excel_bytes():
    """Create a sample Excel file as bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Шаблон"

    # Add headers
    ws.append(["Header1", "Header2", "Header3"])
    ws.append(["Артикул *", "Название", "Бренд в одежде и обуви"])

    # Add data
    ws.append(["ART001", "Товар 1", "Shuzzi"])
    ws.append(["ART002", "Товар 2", "OtherBrand"])
    ws.append(["ART003", "Товар 3", "Shuzzi"])

    # Add video sheet
    ws_video = wb.create_sheet("Озон.Видео")
    ws_video.append(["Header"])
    ws_video.append(["Артикул *", "Видео URL"])
    ws_video.append(["ART001", "http://example.com/video1"])
    ws_video.append(["ART003", "http://example.com/video3"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


class TestFindColumnIndex:
    """Tests for find_column_index function."""

    def test_find_existing_column(self, sample_dataframe):
        """Test finding an existing column."""
        idx = find_column_index(sample_dataframe, "Артикул", search_row=1)
        assert idx == 0

    def test_find_brand_column(self, sample_dataframe):
        """Test finding brand column."""
        idx = find_column_index(sample_dataframe, "Бренд в одежде и обуви", search_row=1)
        assert idx == 2

    def test_column_not_found(self, sample_dataframe):
        """Test when column doesn't exist."""
        idx = find_column_index(sample_dataframe, "NonExistent", search_row=1)
        assert idx is None

    def test_empty_dataframe(self):
        """Test with empty DataFrame."""
        df = pd.DataFrame()
        idx = find_column_index(df, "Any", search_row=0)
        assert idx is None


class TestFilterByBrand:
    """Tests for filter_by_brand function."""

    def test_filter_shuzzi(self, sample_dataframe):
        """Test filtering for Shuzzi brand."""
        filtered = filter_by_brand(sample_dataframe, "Shuzzi", brand_column_index=2, header_rows=2)

        # Should have 2 header rows + 2 data rows with Shuzzi
        assert len(filtered) == 4
        # Check that non-Shuzzi row is filtered out
        values = filtered.iloc[2:, 2].tolist()
        assert all("Shuzzi" in str(v) for v in values)

    def test_filter_case_insensitive(self, sample_dataframe):
        """Test that filtering is case-insensitive."""
        filtered = filter_by_brand(sample_dataframe, "shuzzi", brand_column_index=2, header_rows=2)
        assert len(filtered) == 4

    def test_filter_no_matches(self, sample_dataframe):
        """Test filtering with no matches."""
        filtered = filter_by_brand(sample_dataframe, "NonExistentBrand", brand_column_index=2, header_rows=2)
        # Should return only headers
        assert len(filtered) == 2

    def test_filter_invalid_column(self, sample_dataframe):
        """Test with invalid column index."""
        filtered = filter_by_brand(sample_dataframe, "Shuzzi", brand_column_index=999, header_rows=2)
        # Should return original dataframe
        assert len(filtered) == len(sample_dataframe)


class TestExtractArticleCodes:
    """Tests for extract_article_codes function."""

    def test_extract_from_template(self, sample_excel_bytes):
        """Test extracting article codes from template sheet."""
        codes = extract_article_codes(sample_excel_bytes, "Шаблон")

        assert len(codes) == 3
        assert "ART001" in codes
        assert "ART002" in codes
        assert "ART003" in codes

    def test_extract_from_nonexistent_sheet(self, sample_excel_bytes):
        """Test extracting from non-existent sheet."""
        codes = extract_article_codes(sample_excel_bytes, "NonExistent")
        assert len(codes) == 0


class TestFilterByArticles:
    """Tests for filter_by_articles function."""

    def test_filter_by_article_set(self, sample_dataframe):
        """Test filtering by article codes."""
        article_codes = {"ART001", "ART003"}
        filtered = filter_by_articles(sample_dataframe, article_codes, header_rows=2)

        # Should have 2 header rows + 2 matching data rows
        assert len(filtered) == 4

    def test_filter_empty_set(self, sample_dataframe):
        """Test filtering with empty article set."""
        filtered = filter_by_articles(sample_dataframe, set(), header_rows=2)
        # Should return original
        assert len(filtered) == len(sample_dataframe)

    def test_filter_no_matches(self, sample_dataframe):
        """Test filtering with no matching articles."""
        article_codes = {"NONEXISTENT"}
        filtered = filter_by_articles(sample_dataframe, article_codes, header_rows=2)
        # Should return only headers
        assert len(filtered) == 2


class TestGetSheetInfo:
    """Tests for get_sheet_info function."""

    def test_get_info(self, sample_excel_bytes):
        """Test getting sheet information."""
        info = get_sheet_info(sample_excel_bytes)

        assert len(info) == 2
        assert info[0][0] == "Шаблон"
        assert info[0][1] == 5  # 2 header + 3 data rows
        assert info[1][0] == "Озон.Видео"


class TestReadExcelSheet:
    """Tests for read_excel_sheet function."""

    def test_read_without_filters(self, sample_excel_bytes):
        """Test reading sheet without filters."""
        config = MergeConfig(
            sheets={"Шаблон": SheetConfig(name="Шаблон", include=True, header_rows=2)},
        )
        sheet_config = config.sheets["Шаблон"]

        df = read_excel_sheet(sample_excel_bytes, "Шаблон", config, sheet_config)

        assert not df.empty
        assert len(df) == 5  # 2 headers + 3 data

    def test_read_with_brand_filter(self, sample_excel_bytes):
        """Test reading sheet with brand filter."""
        config = MergeConfig(
            sheets={
                "Шаблон": SheetConfig(
                    name="Шаблон",
                    include=True,
                    header_rows=2,
                    filter_by_brand=True,
                )
            },
            brand_filter="Shuzzi",
        )
        sheet_config = config.sheets["Шаблон"]

        df = read_excel_sheet(sample_excel_bytes, "Шаблон", config, sheet_config)

        assert not df.empty
        # Should have 2 headers + 2 Shuzzi rows
        assert len(df) == 4

    def test_read_with_article_filter(self, sample_excel_bytes):
        """Test reading sheet with article filter."""
        config = MergeConfig(
            sheets={
                "Озон.Видео": SheetConfig(
                    name="Озон.Видео",
                    include=True,
                    header_rows=2,
                    filter_by_articles=True,
                )
            },
        )
        sheet_config = config.sheets["Озон.Видео"]

        # Only keep ART001
        template_articles = {"ART001"}

        df = read_excel_sheet(
            sample_excel_bytes,
            "Озон.Видео",
            config,
            sheet_config,
            template_articles=template_articles,
        )

        assert not df.empty
        # Should have 2 headers + 1 matching row
        assert len(df) == 3


class TestMergeExcelFiles:
    """Tests for merge_excel_files function."""

    def test_merge_single_file(self, sample_excel_bytes):
        """Test merge with single file (no additional files)."""
        config = MergeConfig(
            sheets={
                "Шаблон": SheetConfig(name="Шаблон", include=True, header_rows=2),
            },
        )

        result_bytes = merge_excel_files(sample_excel_bytes, [], config)

        assert result_bytes is not None
        assert len(result_bytes) > 0

        # Verify result
        wb = openpyxl.load_workbook(io.BytesIO(result_bytes))
        assert "Шаблон" in wb.sheetnames

    def test_merge_with_brand_filter(self, sample_excel_bytes):
        """Test merge with brand filter applied."""
        config = MergeConfig(
            sheets={
                "Шаблон": SheetConfig(
                    name="Шаблон",
                    include=True,
                    header_rows=2,
                    filter_by_brand=True,
                ),
            },
            brand_filter="Shuzzi",
        )

        result_bytes = merge_excel_files(sample_excel_bytes, [], config)

        # Verify result has filtered data
        wb = openpyxl.load_workbook(io.BytesIO(result_bytes))
        ws = wb["Шаблон"]

        # Should have 2 header + 2 Shuzzi rows
        assert ws.max_row == 4

    def test_merge_multiple_files(self, sample_excel_bytes):
        """Test merging multiple files."""
        config = MergeConfig(
            sheets={
                "Шаблон": SheetConfig(name="Шаблон", include=True, header_rows=2),
            },
        )

        # Use same file as additional (simulating merge)
        result_bytes = merge_excel_files(
            sample_excel_bytes,
            [sample_excel_bytes],
            config,
        )

        # Verify result
        wb = openpyxl.load_workbook(io.BytesIO(result_bytes))
        ws = wb["Шаблон"]

        # Should have 2 header + 3 original + 3 additional = 8 rows
        assert ws.max_row == 8

    def test_merge_with_video_sheet_article_filter(self, sample_excel_bytes):
        """Test merge with video sheet filtered by template articles."""
        config = MergeConfig(
            sheets={
                "Шаблон": SheetConfig(
                    name="Шаблон",
                    include=True,
                    header_rows=2,
                    filter_by_brand=True,
                ),
                "Озон.Видео": SheetConfig(
                    name="Озон.Видео",
                    include=True,
                    header_rows=2,
                    filter_by_articles=True,
                ),
            },
            brand_filter="Shuzzi",
        )

        result_bytes = merge_excel_files(sample_excel_bytes, [], config)

        # Verify
        wb = openpyxl.load_workbook(io.BytesIO(result_bytes))

        # Template should have only Shuzzi rows (ART001, ART003)
        ws_template = wb["Шаблон"]
        assert ws_template.max_row == 4  # 2 headers + 2 Shuzzi

        # Video sheet should have only rows matching Shuzzi articles
        ws_video = wb["Озон.Видео"]
        assert ws_video.max_row == 4  # 2 headers + 2 matching (ART001, ART003)

    def test_merge_excludes_unchecked_sheets(self, sample_excel_bytes):
        """Test that unchecked sheets remain in the result unchanged."""
        config = MergeConfig(
            sheets={
                "Шаблон": SheetConfig(name="Шаблон", include=True, header_rows=2),
                "Озон.Видео": SheetConfig(name="Озон.Видео", include=False, header_rows=2),
            },
        )

        result_bytes = merge_excel_files(sample_excel_bytes, [], config)

        # Both sheets should exist
        wb = openpyxl.load_workbook(io.BytesIO(result_bytes))
        assert "Шаблон" in wb.sheetnames
        assert "Озон.Видео" in wb.sheetnames
        
        # Verify that excluded sheet remained unchanged (4 rows: 2 header + 2 data)
        ws_video = wb["Озон.Видео"]
        assert ws_video.max_row == 4  # Should be original size

    def test_merge_multiple_files_preserves_excluded_sheets(self, sample_excel_bytes):
        """Test that excluded sheets remain unchanged when merging multiple files."""
        config = MergeConfig(
            sheets={
                "Шаблон": SheetConfig(name="Шаблон", include=True, header_rows=2),
                "Озон.Видео": SheetConfig(name="Озон.Видео", include=False, header_rows=2),
            },
        )

        # Merge with additional file
        result_bytes = merge_excel_files(
            sample_excel_bytes,
            [sample_excel_bytes],  # Additional file
            config,
        )

        wb = openpyxl.load_workbook(io.BytesIO(result_bytes))
        
        # Шаблон should be merged (doubled)
        ws_template = wb["Шаблон"]
        assert ws_template.max_row == 8  # 2 header + 3 original + 3 additional

        # Озон.Видео should remain unchanged (not merged)
        ws_video = wb["Озон.Видео"]
        assert ws_video.max_row == 4  # Should be original size, not doubled

    def test_merge_with_progress_callback(self, sample_excel_bytes):
        """Test merge with progress callback."""
        progress_updates = []

        def callback(pct: float, msg: str):
            progress_updates.append((pct, msg))

        config = MergeConfig(
            sheets={
                "Шаблон": SheetConfig(name="Шаблон", include=True, header_rows=2),
            },
        )

        result_bytes = merge_excel_files(sample_excel_bytes, [], config, progress_callback=callback)

        assert result_bytes is not None
        assert len(progress_updates) > 0
        # Check that progress goes from 0 to 1
        assert progress_updates[0][0] > 0
        assert progress_updates[-1][0] == 1.0


class TestExcelMergeError:
    """Tests for error handling."""

    def test_invalid_file_raises_error(self):
        """Test that invalid file raises ExcelMergeError."""
        config = MergeConfig(
            sheets={
                "Шаблон": SheetConfig(name="Шаблон", include=True, header_rows=2),
            },
        )

        with pytest.raises(ExcelMergeError):
            merge_excel_files(b"invalid data", [], config)
